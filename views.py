from datetime import datetime
from flask import Flask, render_template, request, jsonify

import excel2sbol.converter as conv
import sbol2
#import tempfile
import requests
import os
from excel2flapjack.main import X2F
from pandas import ExcelFile
import json

#imports for data extraction
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import re
import math
from flask_cors import CORS

app = Flask(__name__)

CORS(app, resources={r"/*": {"origins": "*"}})

class XDC:

    """XDC class to upload excel file to SynBioHub and Flapjack.

    ...

    Attributes
    ----------
    input_excel_path : str
        path to the input excel file
    fj_url : str
        URL of the Flapjack instance
    fj_user : str
        username of the Flapjack instance
    fj_pass : str
        password of the Flapjack instance
    sbh_url : str
        URL of the SynBioHub instance
    sbh_user : str
        username of the SynBioHub instance
    sbh_pass : str
        password of the SynBioHub instance
    sbh_collection : str
        collection to upload the SBOL file to
    sbh_collection_description : str
        description of the collection
    sbh_overwrite : bool
        whether to overwrite the SBOL file if it already exists
    fj_overwrite : bool
        whether to overwrite the Flapjack project if it already exists  
    fj_token : str
        token to authenticate with Flapjack
    sbh_token : str
        token to authenticate with SynBioHub
    status : str
        status of the process
    
    Methods
    -------
    initialize()
        Initializes the X2F object
    log_in_fj()
        Logs into Flapjack
    log_in_sbh()
        Logs into SynBioHub
    convert_to_sbol()
        Converts the input excel file to SBOL format
    upload_to_fj()
        Uploads the SBOL file to Flapjack
    upload_to_sbh()
        Uploads the SBOL file to SynBioHub
    run()
        Runs the entire process
    """
    def __init__(self, input_excel_path, fj_url, fj_user, fj_pass, sbh_url, sbh_user, sbh_pass, sbh_collection, sbh_collection_description, sbh_overwrite, fj_overwrite, fj_token, sbh_token):
        self.input_excel_path = input_excel_path
        self.fj_url = fj_url
        self.fj_user = fj_user
        self.fj_pass = fj_pass
        self.sbh_url = sbh_url
        self.sbh_user = sbh_user
        self.sbh_pass = sbh_pass
        self.sbh_collection = sbh_collection
        self.sbh_collection_description = sbh_collection_description
        self.sbh_overwrite = sbh_overwrite
        self.fj_overwrite = fj_overwrite
        self.fj_token = fj_token
        self.sbh_token = sbh_token
        #self.status = "Not started"
        self.input_excel = ExcelFile(self.input_excel_path)
        self.xdc = None
        self.homespace = 'https://sbolstandard.org'
        self.sbol_doc = None
        self.sbol_fj_doc = None
        self.sbol_graph_uri = None
        self.file_path_out = f'{sbh_collection}_converted_SBOL.xml'
        self.file_path_out2 = f'{sbh_collection}_SBOL_Fj_doc.xml'


    def initialize(self):
        self.xdc = X2F(excel_path=self.input_excel_path, 
                    fj_url=self.fj_url, 
                    fj_user=self.fj_user, 
                    fj_pass=self.fj_pass, 
                    #fj_token=self.fj_token, #TODO
                    #overwrite=self.fj_overwrite)
        )
        if self.sbh_collection_description is None:
            self.sbh_collection_description = 'Collection of SBOL files uploaded from XDC'
        #self.status = "Initialized"

        
    def log_in_fj(self):
        if not self.fj_token:
            self.xdc.fj_login(username=self.fj_user, password=self.fj_pass)
            self.status = "Logged into Flapjack"
        else:
            pass #TODO
        self.status = "Logged into Flapjack"

    def log_in_sbh(self):
        # SBH Login
        if self.sbh_token is None:
            url = f'{self.sbh_url}/login'
            print(f'Logging in to SBH with URL: {url}')  # Debug statement
            response = requests.post(
                url,
                headers={'Accept': 'text/plain'},
                data={
                    'email': self.sbh_user,
                    'password': self.sbh_pass,
                }
            )
            self.sbh_token = response.text
        #self.status = "Logged into SynBioHub"

    def convert_to_sbol(self):
        #temp_dir = tempfile.TemporaryDirectory() #TODO:check if I need to create the temporary object in a different context
        #file_path_out = os.path.join(temp_dir.name, 'converted_SBOL.xml')
        
        conv.converter(file_path_in = self.input_excel_path, 
                file_path_out = self.file_path_out)
        # Pull graph uri from synbiohub
        response = requests.get(
            f'{self.sbh_url}/profile',
            headers={
                'Accept': 'text/plain',
                'X-authorization': self.sbh_token
                }
        )
        self.sbol_graph_uri = response.json()['graphUri']
        sbol_collec_url = f'{self.sbol_graph_uri}/{self.sbh_collection}/'

        # Parse sbol to create hashmap of flapjack id to sbol uri
        doc = sbol2.Document()
        doc.read(self.file_path_out)
        sbol_hash_map = {}
        for tl in doc:
            if 'https://flapjack.rudge-lab.org/ID' in tl.properties:
                sbol_uri = tl.properties['http://sbols.org/v2#persistentIdentity'][0]
                sbol_uri = sbol_uri.replace(self.homespace, sbol_collec_url)
                sbol_uri = f'{sbol_uri}/1'

                sbol_name = str(tl.properties['http://sbols.org/v2#displayId'][0])
                sbol_hash_map[sbol_name] = sbol_uri
        self.xdc.sbol_hash_map = sbol_hash_map
        self.sbol_doc = doc
        self.status = "Converted to SBOL"

    def upload_to_fj(self):
        self.xdc.upload_all()
        #self.status = "Uploaded to Flapjack"

    def upload_to_sbh(self):
        #temp_dir = tempfile.TemporaryDirectory() #TODO:check if I need to create the temporary object in a different context
        #file_path_out2 = os.path.join(temp_dir.name, 'SBOL_Fj_doc.xml')
        
        # Add flapjack annotations to the SBOL
        doc = sbol2.Document()
        doc.read(self.file_path_out)
        for tl in self.sbol_doc:
            id = str(tl).split('/')[-2]
            if id in self.xdc.sbol_hash_map:
                setattr(tl, 'flapjack_ID',
                        sbol2.URIProperty(tl,
                        'https://flapjack.rudge-lab.org/ID',
                            '0', '*', [], initial_value=f'http://wwww.flapjack.com/{self.xdc.sbol_hash_map[id]}'))
        #doc = sbol2.Document()
        doc.write(self.file_path_out2)

        if self.sbh_overwrite:
            sbh_overwrite = '1'
        else:
            sbh_overwrite = '0'
        # SBH file upload
        response = requests.post(
            f'{self.sbh_url}/submit',
            headers={
                'Accept': 'text/plain',
                'X-authorization': self.sbh_token
            },
            files={
            'files': open(self.file_path_out2,'rb'),
            },
            data={
                'id': self.sbh_collection,
                'version' : '1',
                'name' : self.sbh_collection,
                'description' : self.sbh_collection_description, #TODO
                'overwrite_merge' : sbh_overwrite
            },

        )

        if response.text == "Submission id and version already in use":
            print('not submitted')
            raise AttributeError(f'The collection ({self.sbh_collection}) could not be submitted to synbiohub as the collection already exists and overite is not on.')
        # if response.text == "Successfully uploaded":
        #      success = True
        #self.status = "Uploaded to SynBioHub"
        return f'{self.sbol_graph_uri}/{self.sbh_collection}/{self.sbh_collection}_collection/1'

        

    def run(self):
        self.initialize()
        #self.log_in_fj()
        self.convert_to_sbol()
        #self.upload_to_fj()
        self.upload_to_sbh()



#routes

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/about/")
def about():
    return render_template("about.html")

@app.route("/contact/")
def contact():
    return render_template("contact.html")

@app.route("/hello/")
@app.route("/hello/<name>")
def hello_there(name = None):
    return render_template(
        "hello_there.html",
        name=name,
        date=datetime.now()
    )

@app.route("/api/data")
def get_data():
    return app.send_static_file("data.json")

@app.route('/uploader', methods = ['POST'])
def upload_file():
    if request.method == 'POST':
        f = request.files['file']
        fj_user = request.form['fjusername']
        fj_pass = request.form['fjpwd']
        sbh_user = request.form['sbhusername']
        sbh_pass = request.form['sbhpwd']
        sbh_collec = request.form['sbhcollec']

        if 'sbhover' in request.form:
            sbh_overwrite = True
        else:
            sbh_overwrite = False

        try:
            sbol_collec_url = xdc.experimental_data_uploader(f, fj_user, fj_pass,
                                    sbh_user, sbh_pass, sbh_collec, sbh_overwrite=sbh_overwrite,
                                    fj_overwrite=True)
            sbol_collec_url = f'{sbol_collec_url}{sbh_collec}_collection/1'
            return render_template('upload_success.html', collec_uploaded=sbol_collec_url)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            lnum = exc_tb.tb_lineno
            ex = f'Exception is: {e}, exc_type: {exc_type}, exc_obj: {exc_obj}, fname: {fname}, line_number: {lnum}, traceback: {traceback.format_exc()}'
            return render_template('upload_failure.html', collec_uploaded=sbh_collec, error_message=ex)

@app.route('/test_uploader')
def test_upload_file():
    # params
    fj_url = "localhost:8000" #local
    #fj_url = "flapjack.rudge-lab.org:8000" #Web Instance Rudge Lab
    #fj_url = "198.59.83.73:8000" #Web Instance Genetic Logic Lab
    fj_user = "Gonza10V"
    fj_pass = "010101"

    sbh_url = "https://synbiohub.colorado.edu"
    sbh_user = "Gonza10V"
    sbh_pass = "010101"
    sbh_collec = "xdc_sbs_revamp"
    # updated xcel file is in test/test_files
    test_file_path = '/./static'
    #excel_path = os.path.join(test_file_path, 'flapjack_excel_converter_revamp_medias.xlsx')#"flapjack_excel_converter_revamp2_test.xlsx")
    print(os.getcwd())
    #xcel_path = os.path.join(test_file_path, 'SBS_XDC_template_test.xlsx')#"flapjack_excel_converter_revamp2_test.xlsx")
    excel_path = '/Users/gonzalovidal/Documents/GitHub/python-sample-vscode-flask-tutorial/hello_app/static/SBS_XDC_template_test.xlsx'


    fj_overwrite = False
    sbh_overwrite=False

    # instantiate the XDC class
    xdc = XDC(input_excel_path = excel_path,
            fj_url = fj_url,
            fj_user = fj_user, 
            fj_pass = fj_pass, 
            sbh_url = sbh_url, 
            sbh_user = sbh_user, 
            sbh_pass = sbh_pass, 
            sbh_collection = sbh_collec, 
            sbh_collection_description = 'Default Collection description of SBOL files uploaded from XDC',
            sbh_overwrite = sbh_overwrite, 
            fj_overwrite = fj_overwrite, 
            fj_token = None, 
            sbh_token = None)
    
    xdc.initialize()
    xdc.log_in_sbh()
    xdc.convert_to_sbol()
    sbh_url = xdc.upload_to_sbh()
    return sbh_url



@app.route('/upload_sbs', methods = ['POST'])
def upload_file_from_sbs_post():
    if 'Metadata' not in request.files:
        print(request)
        return 'No file part', 400
    file = request.files['Metadata']
    if file.filename == '':
        return 'No selected file', 400
    file_contents = file.read()
    #if 'Output' not in request.files:
    #    return 'No output file part', 400
    #output_file = request.files['Output']
    #if output_file.filename == '':
    #    return 'No selected output file', 400
    #output_file_contents = output_file.read()
    if 'Params' not in request.files:
        return 'No Params file part', 400
    params_file = request.files['Params']
    if params_file.filename == '':
        return 'No selected Params file', 400
    params_from_request = json.loads(params_file.read())
    auth_token = request.form.get('AuthToken')
    if not auth_token:
        return 'No AuthToken provided', 400

    # instantiate the XDC class using the params_from_request dictionary
    print(request.files['Metadata'])
    xdc = XDC(input_excel_path = request.files['Metadata'],
            fj_url = params_from_request['fj_url'],
            fj_user = params_from_request['fj_user'], 
            fj_pass = params_from_request['fj_pass'], 
            sbh_url = params_from_request['sbh_url'], 
            sbh_user = params_from_request['sbh_user'], 
            sbh_pass = params_from_request['sbh_pass'], 
            sbh_collection = params_from_request['sbh_collec'], 
            sbh_collection_description = 'Default Collection description of SBOL files uploaded from XDC',
            sbh_overwrite = params_from_request['sbh_overwrite'], 
            fj_overwrite = params_from_request['fj_overwrite'], 
            fj_token = None, 
            sbh_token = None)

    xdc.initialize()
    xdc.log_in_sbh()
    xdc.convert_to_sbol()
    sbh_url = xdc.upload_to_sbh()

    sbs_upload_response_dict ={
        "sbh_url": sbh_url,
        "auth_token": auth_token,
        "status": "success"
    }
    return jsonify(sbs_upload_response_dict)



class XDE:

    """XDE (Experimental Data Extractor) class to extract experimental data from
    plate reader excel output and writes it in an XDC template.

    ...

    Attributes
    ----------


    Methods
    -------
    getFileNameFromString(string)
        Extracts the file name from a string
    generateSampleData(file_list,sheet_to_read_from,time_col_name,data_cols_offset)
        Generates sample data from the input excel files
    getNumRows(dataframe,starting_row_idx,starting_col_idx)
        Gets the number of rows for the data
    buildFinalDF(file_list,sample_data_list,time_col_name,data_cols_offset,num_rows_btwn_data,sheet_to_read_from)
        Builds the final dataframe
    writeToMeasurements(XDC_file_name,final_dataframe)
        Writes the final dataframe to the measurements sheet

    """
    def __init__(self):
        pass

    def getFileNameFromString(string):
        pattern = '[\w-]+?(?=\.)'
        # searching the pattern
        result = re.search(pattern, string)
    
        return result.group()

    def generateSampleData(self, file_list, sheet_to_read_from, time_col_name, data_cols_offset): 
        num_assays = len(file_list)
        file_name_list = []
        print(file_list)
        # process experimental data files
        #for i in range(num_assays):
        #    print('generating sample data from' + file_list[i] + 'from a total of' + str(num_assays))
        #    file_name_list.append(self.getFileNameFromString(string=file_list[0])) #TODO correct for multiple files
        file_name_list = ['240223FL50h']


        #final products
        result = pd.DataFrame()
        sample_data_list = []

        #components:result
        assay_id = []
        column = []
        row = []
        sample_id = []

        #componenets:sample_data_list
        columnID = []
        assay_num = []

        #processing:main
        current_sample_id = 1

        for i in range(num_assays):

            current_num_assay = i + 1

            #locating instances of time_col_name
            raw_df = pd.read_excel(file_list[0],sheet_to_read_from) #TODO modify for multiple data outputs
            rows, cols = np.where(raw_df == time_col_name)
            time_col_locations = list(zip(rows, cols))
            num_rows = self.getNumRows(raw_df,rows[0],cols[0])
            
            #extracting signal 1 data to check for blank columns
            start_row = time_col_locations[0][0] + 1
            start_col = data_cols_offset
            num_cols = 96
            working_df = raw_df.iloc[start_row:start_row + num_rows, start_col:start_col + num_cols] #maybe subtract 1 from num rows

            # Check for completely blank (all NaN) columns using numpy
            is_blank = working_df.isna().all().to_numpy()

            # Get the indices of non-blank columns
            data_col_IDX = np.where(~is_blank)[0]
            
            #add to lists
            for j in range(len(data_col_IDX)):
                #result
                assay_id.append(file_name_list[i])
                column.append(data_col_IDX[j] % 12 + 1) #IDX % 12 + 1
                row.append((data_col_IDX[j]//12) + 1)   #IDX // 12 + 1
                sample_id.append(f"Sample{current_sample_id}")
                current_sample_id += 1

                #sample_data_list
                columnID.append(data_col_IDX[j])
                assay_num.append(current_num_assay)

        #assembly:result
        result.insert(0, "Assay ID", assay_id)
        result.insert(0, "Column", column)
        result.insert(0, "Row", row)
        result.insert(0, "Sample ID", sample_id)
        
        #assembly:sample_data_list
        for i in range(len(result)):
            temp_tuple = (sample_id[i],columnID[i],assay_num[i])
            sample_data_list.append(temp_tuple)
        
        with pd.ExcelWriter(file_list[0], mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                result.to_excel(writer,'Sample',index=False)

        return sample_data_list

    def getNumRows(self, dataframe, starting_row_idx, starting_col_idx):
        num_rows = 0
        counter = 1
        time_col_value = dataframe.iloc[starting_row_idx, starting_col_idx]

        while True:
            current_cell = dataframe.iloc[starting_row_idx + counter, starting_col_idx]
            if pd.isna(current_cell) or current_cell == time_col_value:
                break
            if(len(dataframe) <= counter + starting_row_idx + 1): #edge case for if there is only one signal, IDK why i have to add a +1
                num_rows += 1
                break

            counter += 1
            num_rows += 1
            
        return num_rows 

    def buildFinalDF(self, file_list, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data, sheet_to_read_from):
        print(file_list)
        output = pd.DataFrame() 
        time_col_locations = []
        num_rows_per_assay = []
        dataframe_list = []
        num_assays = len(file_list) - 1

        for i in range(num_assays):
            raw_df = pd.read_excel(file_list[i+1],sheet_to_read_from)
            rows, cols = np.where(raw_df == time_col_name)
            temp = list(zip(rows, cols))
            num_rows_per_assay.append(self.getNumRows(raw_df,rows[0],cols[0]))
            time_col_locations.append(temp)
            dataframe_list.append(pd.read_excel(file_list[i + 1],sheet_to_read_from))
            
        for i in range(len(sample_data_list)):  #initilizing information about the current sample and its results #TODO modifiying for only one file, redo for multiple
            rows_to_be_read = []
            current_sample_id = str(sample_data_list[i-1][0])               
            current_col = sample_data_list[i-1][1]
            current_assay = sample_data_list[i-1][2]
            current_first_row = time_col_locations[current_assay-1][0][0] + 1 #[current_assay-1][0][0] + 1
            current_time_col = time_col_locations[current_assay-1][0][1]
            current_num_rows = num_rows_per_assay[current_assay-1]
            current_num_signals = len(time_col_locations[current_assay-1])

            for j in range(current_num_signals): 
                rows_to_be_read.extend(list(range(current_first_row + ((current_num_rows + num_rows_btwn_data + 1)* j), current_first_row + current_num_rows + ((current_num_rows + num_rows_btwn_data + 1)* j))))
            working_df = dataframe_list[current_assay - 1].iloc[rows_to_be_read,[current_time_col,current_col + data_cols_offset]].copy() # at this point it will be the time col and current col for both signals
            working_df.columns = ["Time", "Value"]
            #add signal label
            signal_id = []
            for k in range(current_num_signals):
                signal_id.extend([f"Signal{k + 1}"] * current_num_rows)
            working_df.insert(0, "Signal ID", signal_id)

            #add sample label
            sample_id = [current_sample_id] * len(working_df)
            working_df.insert(0, "Sample ID", sample_id)

            #concat working_df and output
            output = pd.concat([output, working_df], ignore_index=True)

        #add measurement
        measurement_id = []
        for i in range(len(output)):
            measurement_id.append(f"Measurement{i}")
        output.insert(0, "Measurement ID", measurement_id)

        return output

    def buildFinalDFCSV(self, file_list, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data):
        output = pd.DataFrame() 
        time_col_locations = []
        num_rows_per_assay = []
        dataframe_list = []
        num_assays = len(file_list) - 1

        for i in range(num_assays):
            raw_df = pd.read_csv(file_list[i+1])
            rows, cols = np.where(raw_df == time_col_name)
            temp = list(zip(rows, cols))
            num_rows_per_assay.append(self.getNumRows(raw_df,rows[0],cols[0]))
            time_col_locations.append(temp)
            dataframe_list.append(pd.read_csv(file_list[i + 1]))
            
        for i in range(len(sample_data_list)):  #initilizing information about the current sample and its results
            rows_to_be_read = []
            current_sample_id = sample_data_list[i][0]               
            current_col = sample_data_list[i][1]
            current_assay = sample_data_list[i][2]
            current_first_row = time_col_locations[current_assay - 1][0][0] + 1
            current_time_col = time_col_locations[current_assay - 1][1][1]
            current_num_rows = num_rows_per_assay[current_assay - 1]
            current_num_signals = len(time_col_locations[current_assay - 1])

            for j in range(current_num_signals): 
                rows_to_be_read.extend(list(range(current_first_row + ((current_num_rows + num_rows_btwn_data + 1)* j), current_first_row + current_num_rows + ((current_num_rows + num_rows_btwn_data + 1)* j))))
            
            working_df = dataframe_list[current_assay - 1].iloc[rows_to_be_read,[current_time_col,current_col + data_cols_offset]].copy() # at this point it will be the time col and current col for both signals
            working_df.columns = ["Time", "Value"]

            #add signal label
            signal_id = []
            for k in range(current_num_signals):
                signal_id.extend([f"Signal{k + 1}"] * current_num_rows)
            working_df.insert(0, "Signal ID", signal_id)

            #add sample label
            sample_id = [current_sample_id] * len(working_df)
            working_df.insert(0, "Sample ID", sample_id)

            #concat working_df and output
            output = pd.concat([output, working_df], ignore_index=True)

        #add measurement
        measurement_id = []
        for i in range(len(output)):
            measurement_id.append(f"Measurement{i}")
        output.insert(0, "Measurement ID", measurement_id)

        return output


    def writeToMeasurements(self, XDC_file_name, final_dataframe):
        book = load_workbook(XDC_file_name)
        sheet = book['Measurement']

        # Clear the existing data in the 'Measurement' sheet
        sheet.delete_rows(1, sheet.max_row)

        # Write the headers
        sheet.append(['Measurement ID', 'Sample ID', 'Signal ID', 'Time', 'Value'])

        # Write the data
        for row in final_dataframe.itertuples(index=False):
            sheet.append(list(row))

        book.save(XDC_file_name)
        book.close()

        return
    
#create xde object
xde = XDE()
   
upload_dir = os.path.join(os.getcwd(), 'uploads')
if not os.path.exists(upload_dir):
    os.makedirs(upload_dir)

def process_files_with_script(uploaded_files, metadata_file, preset):
    # Save the uploaded files to disk
    print('process started')
    file_paths = []
    # get metadata file path
    metadata_file_path = os.path.join('uploads', metadata_file.filename)
    for file in uploaded_files:
        print(str(file.filename))
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)
        print(str(file_path))
        file_paths.append(file_path)
    
    # Determine the parameters based on the selected preset
    if preset == 'spark_10m':
        time_col_name = "Time [s]"
        data_cols_offset = 3
        num_rows_btwn_data = 2
        file_type_choice = 1
        sheet_to_read_from = 0
    elif preset == 'synergy_h1':
        time_col_name = "Time"
        data_cols_offset = 2
        num_rows_btwn_data = 3
        file_type_choice = 2
        sheet_to_read_from = 0

    elif preset == 'chan_lab':
        time_col_name = "UNIX Timestamp"
        data_cols_offset = 2
        num_rows_btwn_data = 0
        file_type_choice = 2
        sheet_to_read_from = 0

    elif preset == 'CLAIROstar':
        time_col_name = "Time"
        data_cols_offset = 2
        num_rows_btwn_data = 0
        file_type_choice = 1
        sheet_to_read_from = 0 # TODO this was 1 for some reason before, review
    else:
        # Handle manual configuration if needed
        pass
    print(time_col_name)

    # Call the appropriate function based on the file type
    sample_data_list = xde.generateSampleData(file_list=file_paths, sheet_to_read_from=sheet_to_read_from, time_col_name=time_col_name, data_cols_offset=data_cols_offset)
    print('sample_data_list generated')
    if file_type_choice == 1:
        final_df = xde.buildFinalDF(file_paths, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data, sheet_to_read_from)
        print('final_df generated from xlsx')
        final_df.to_csv('uploads/final_df.csv', index=False)
    else:
        final_df = xde.buildFinalDFCSV(file_paths, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data)
        print('final_df generated form csv')

    # Write the modified file to disk
    modified_file_path = os.path.join('uploads', 'modified_file.xlsx')
    xde.writeToMeasurements(file_paths[0], final_df)
    print('final_df written to measurements')

    # Return the modified file path
    return file_paths[0]
    
@app.route('/extract_data', methods=['POST'])
def process_files():
    if 'Metadata' not in request.files:
        print(request)
        return 'No file part', 400
    metadata_file = request.files['Metadata']
    if metadata_file.filename == '':
        return 'No selected file', 400

    if 'platereader_output' not in request.files:
        print(request)
        return 'No file part', 400
    file = request.files['platereader_output']
    if file.filename == '':
        return 'No selected file', 400
    # Get uploaded files from request
    platereader_output_files = [file] #request.files.getlist('platereader_output') #[file] #TODO: see how to iterate through files 
    # Get selected preset from request
    preset = request.form.get('preset')

    print(platereader_output_files)
    print(preset)
    # Call the process_files_with_script function
    modified_file_path = process_files_with_script(platereader_output_files, metadata_file, preset)

    # Return the modified file as a response
    with open(modified_file_path, 'rb') as f:
        modified_file_data = f.read()
    #os.remove(modified_file_path)  # Optional: Remove the temporary file
    print('finished')
    # save modified file to disk



    return modified_file_data