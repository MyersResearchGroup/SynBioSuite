from __future__ import annotations
from pydoc import doc 
from flask import request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from .main import app
from .utils import abstract_design_2_plasmids, sbol2build_moclo 
from .version import __version__
import sys
import os
import json
import xml.etree.ElementTree as ET
import tricahue
import sbol2 as sb2
import pudu
import subprocess
import requests
import sbol2
import sbol2.model
import excel2sbol
from uuid import uuid4
from urllib.parse import urlencode
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re

#routes
#check if the app is running
@app.route('/api/status')
def pin():
    return jsonify({"status": "working", "version": __version__}), 200

@app.route('/api/downloadTemplate', methods = ['POST'])
def download_template():
    return sbh_download_template(request.files)

@app.route('/api/uploadSBOL', methods = ['POST'])
def upload_sbol():
    return sbh_upload(request.files)

@app.route('/api/uploadResource', methods = ['POST'])
def upload_resource():
    return sbh_fj_upload(request.files)

@app.route('/api/uploadAssembly', methods = ['POST'])
def upload_assembly():
    return 'Not implemented yet', 501

@app.route('/api/uploadTransformation', methods = ['POST'])
def upload_transformation():
    return 'Not implemented yet', 501

@app.route('/api/uploadExperiment', methods = ['POST'])
def upload_experiment():
    return sbh_fj_upload(request.files)

def make_identifier(text: str) -> str:
    """
    Convert arbitrary text into a valid SBOL displayId.

    - Replaces non-alphanumeric characters with '_'
    - Collapses multiple underscores
    - Removes leading/trailing underscores
    - Ensures the first character is a letter or '_'
    """

    identifier = text.strip()

    # Replace non-alphanumeric characters with underscores
    identifier = re.sub(r'[^A-Za-z0-9]+', '_', identifier)

    # Collapse multiple underscores
    identifier = re.sub(r'_+', '_', identifier)

    # Remove leading/trailing underscores
    identifier = identifier.strip('_')

    # Ensure non-empty
    if not identifier:
        identifier = "_"

    # SBOL displayIds must start with a letter or underscore
    elif not re.match(r'^[A-Za-z_]', identifier):
        identifier = "_" + identifier

    return identifier

'''
Helper function to perform SPARQL queries to fetch URIs from SynBioHub
'''
def sbh_get_subCollection_uris(sbh_url, sbh_token, usergraph, collectionUri, role = None, type = None):
    if role is None and type is None:
        query = f'PREFIX sbol: <http://sbols.org/v2#> SELECT ?s ?id FROM <{usergraph}> WHERE {{ <{collectionUri}> sbol:member ?s . ?s sbol:displayId ?id }}'
    elif type is None:
        query = f'PREFIX sbol: <http://sbols.org/v2#> SELECT ?s ?id FROM <{usergraph}> WHERE {{ ?s sbol:role <{role}> . <{collectionUri}> sbol:member ?s . ?s sbol:displayId ?id }}'
    elif role is None:
        query = f'PREFIX sbol: <http://sbols.org/v2#> SELECT ?s ?id FROM <{usergraph}> WHERE {{ ?s sbol:type <{type}> . <{collectionUri}> sbol:member ?s . ?s sbol:displayId ?id }}'
    else:
        query = f'PREFIX sbol: <http://sbols.org/v2#> SELECT ?s ?id FROM <{usergraph}> WHERE {{ ?s sbol:type <{type}> . ?s sbol:role <{role}> . <{collectionUri}> sbol:member ?s . ?s sbol:displayId ?id }}'

    url = f"{sbh_url}/sparql?{urlencode({'query': query})}"
    response =  requests.get(
        url,
        headers={
            'Accept': 'application/json',
            'X-authorization': sbh_token
            },
        )
    if not response.ok:
        raise Exception(f"SynBioHub sparql query failed ({response.status_code}): {response.text}")
    return response.json()

'''
Helper function to download Excel template
'''
def fetch_template_bytes(object_type_id):
    templates = {
        "synbio.object-type.sample-designs": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/SampleDesign.xlsm",
            "SampleDesign.xlsm",
        ),
        "synbio.object-type.strains": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Strains.xlsm",
            "Strains.xlsm",
        ),
        "synbio.object-type.resources": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Resources.xlsm",
            "Resources.xlsm",
        ),
        "synbio.object-type.study-data": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Assay.xlsm",
            "Assay.xlsm",
        ),
    }

    if object_type_id not in templates:
        raise ValueError(f"Unknown object type: {object_type_id}")

    url, filename = templates[object_type_id]
    response = requests.get(url)
    response.raise_for_status()
    return response.content, filename

def expand_table_in_xlsm(wb, sheet_name, table_name, new_rows):
    ws = wb[sheet_name]
    table = ws.tables[table_name]

    start_cell, end_cell = table.ref.split(":")
    start_col_letters = "".join(ch for ch in start_cell if ch.isalpha())
    start_row = int("".join(ch for ch in start_cell if ch.isdigit()))
    end_col_letters = "".join(ch for ch in end_cell if ch.isalpha())
    end_row = int("".join(ch for ch in end_cell if ch.isdigit()))

    start_col = column_index_from_string(start_col_letters)
    end_col = column_index_from_string(end_col_letters)

    # Find the first blank row inside the existing table body
    first_blank_row = None
    for row_idx in range(start_row + 1, end_row + 1):
        if all(ws.cell(row=row_idx, column=col_idx).value is None for col_idx in range(start_col, end_col + 1)):
            first_blank_row = row_idx
            break

    # Write into the blank row first if one exists
    write_row = first_blank_row if first_blank_row is not None else end_row + 1

    for values in new_rows:
        # If we run past the current table range, extend it
        if write_row > end_row:
            end_row = write_row
            table.ref = f"{start_cell}:{end_col_letters}{end_row}"

        for col_idx, value in enumerate(values, start=start_col):
            ws.cell(row=write_row, column=col_idx, value=value)

        write_row += 1

'''
Helper function to download Excel template and populate tables from SynBioHub
'''
def sbh_download_template(files):
    if 'Params' not in files:
        return 'No Params file part', 400

    params_file = files['Params']
    if params_file.filename == '':
        return 'No selected Params file', 400

    params_from_request = json.loads(params_file.read())

    required_params = ['template_type', 'sbh_url', 'sbh_token', 'collection_url']
    for param in required_params:
        if param not in params_from_request:
            return f'Parameter {param} not found in request', 400

    if params_from_request['sbh_token'] is None:
        return 'No SBH credentials provided', 400

    sbh_url = params_from_request['sbh_url']
    sbh_collection_url = params_from_request['collection_url'] 
    sbh_token = params_from_request['sbh_token']
    parts = sbh_collection_url.split("/")
    usergraph = "/".join(parts[:5])

    template_type = params_from_request['template_type']

    try:
        template_bytes, filename = fetch_template_bytes(template_type)
        wb = load_workbook(BytesIO(template_bytes), keep_vba=True)

        if template_type == "synbio.object-type.strains":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/ncit/NCIT:C14419")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_chassis_collections",table_name="SBH_chassis_collections",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/so/SO:0000637")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_plasmids_collections",table_name="SBH_plasmids_collections",new_rows=new_rows)
        elif template_type == "synbio.object-type.sample-designs":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/ncit/NCIT:C48164")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_media_collection",table_name="SBH_media_collection",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,type="http://www.biopax.org/release/biopax-level3.owl#SmallMolecule")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_chemicals_collection",table_name="SBH_chemicals_collection",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://purl.obolibrary.org/obo/NCIT_C97158")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_strains_collection",table_name="SBH_strains_collection",new_rows=new_rows)
        elif template_type == "synbio.object-type.study-data":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"https://wiki.synbiohub.org/wiki/Terms/SynBioSuite#SampleDesign")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_sampledesigns_collection",table_name="SBH_sampledesigns_collection",new_rows=new_rows)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

def upload_sbh_attachments(sbh_token, sbh_collection_url, file, importType, collection_url, experimentId):
    print("uploading attachments")
 
    version = '1'
 
#    parts = sbh_collection_url.split("/")
#    subCollection_url = "/".join(parts[:6]) + "/" + importType + "/1"
#    search_result = self.sbh_get_attachment_uri(self.sbh_url,self.sbh_token,self.sbh_collection_url,attachment_name)
    # for binding in search_result["results"]["bindings"]:
    #     uri = binding["s"]["value"]
    #     print(f"Deleting existing attachment {uri}")
    #     response = requests.get(f'{uri}/remove', headers=headers)
    #     if not response.ok:
    #         raise Exception(f"Deleting existing attachment failed ({response.status_code}): {response.text}")
    headers = {'Accept': 'text/plain', 'X-authorization': sbh_token}
    response = requests.post(f'{collection_url}/attach', headers=headers, files=file)
    if not response.ok:
        raise Exception(f"Uploading attachments to SynBioHub failed ({response.status_code}): {response.text}")
        print(f'Uploaded attachment {upload_file["file"][0]}: {response.status_code}')

'''
Helper function to upload SBOL to SynBioHub
'''
def sbh_upload(files):
    if 'SBOL' in files:
        sbol_file = files['SBOL']
        sbml_file = None
    elif 'SBML' in files:
        sbml_file = files['SBML']
        sbol_file = None
    else:
        return 'No file part', 400
 
    if sbol_file and sbol_file.filename != '':
        root, extension = os.path.splitext(sbol_file.filename)
    elif sbml_file and sbml_file.filename != '':
        root, extension = os.path.splitext(sbml_file.filename)
    else:
        return 'No selected file', 400
    if not extension == '.xml':
        return 'Invalid file format', 400

    # Check params from frontend
    if 'Params' not in files:
        return 'No Params file part', 400
    params_file = files['Params']
    if params_file.filename == '':
        return 'No selected Params file', 400
    params_from_request = json.loads(params_file.read())
    sbh_url = params_from_request.get('sbh_url')
    if sbh_url and not (sbh_url.startswith('http://') or sbh_url.startswith('https://')):
        params_from_request['sbh_url'] = 'https://' + sbh_url

    required_params = ['sbh_url', 'sbh_token', 'collection_url', 'sbh_overwrite']

    for param in required_params:
        if param not in params_from_request:
            return 'Parameter ' + param + ' not found in request', 400
    if (params_from_request['sbh_token'] is None):
        return 'No SBH credentials provided', 400
    sbh_url = params_from_request['sbh_url']
    sbh_collection_url = params_from_request['collection_url'] 
    sbh_overwrite = params_from_request['sbh_overwrite'] 
    sbh_token = params_from_request['sbh_token']
    importType = params_from_request['importType']
    parts = sbh_collection_url.split("/")
    usergraph = "/".join(parts[:5])
    subCollection_url = "/".join(parts[:6]) + "/" + importType + "/1"

    upload_dir = os.path.join(os.getcwd(), "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    if sbol_file and sbol_file.filename != '':
        safe_sbol_filename = secure_filename(sbol_file.filename)
        if safe_sbol_filename == '':
            return 'Invalid SBOL file name', 400
        sbol_path = os.path.join(
            upload_dir,
            f"{uuid4()}_{safe_sbol_filename}"
        )
        sbol_out_path = os.path.join(
            upload_dir,
            f"{uuid4()}_out_{safe_sbol_filename}"
        )
        sbml_path = None
    else:
        safe_sbml_filename = secure_filename(sbml_file.filename)
        if safe_sbml_filename == '':
            return 'Invalid SBML file name', 400
        sbml_path = os.path.join(
            upload_dir,
            f"{uuid4()}_{safe_sbml_filename}"
        )
        sbol_out_path = os.path.join(
            upload_dir,
            f"{uuid4()}_out_{safe_sbml_filename}"
        )
        sbol_path = None

    try:
        if sbol_file and sbol_file.filename != '':
            sbol_file.save(sbol_path)
        else:
            sbml_file.save(sbml_path)
        sbol2.Config.setOption(sbol2.ConfigOptions.SBOL_COMPLIANT_URIS, True)
        sbol2.Config.setOption(sbol2.ConfigOptions.SBOL_TYPED_URIS, False)
        homespaces = {
            "Devices": "https://example.com/",
            "Designs": "https://sbolcanvas.org/",
            "Models": "https://sbolcanvas.org/",
            "Plasmids": "https://example.com/",
        }
        sbol2.setHomespace(homespaces.get(importType, "https://synbiosuite.org/"))
        doc = sbol2.Document()
        if sbol_file and sbol_file.filename != '':
            doc.read(sbol_path)
        else:
            with open(sbml_path, "rb") as fobj:
                response = requests.post(
                    f"{subCollection_url}/attach",
                    headers={
                        "Accept": "text/plain",
                        "X-authorization": sbh_token,
                    },
                    files={
                        "file": (os.path.basename(sbml_path), fobj),
                    },
                )

                if not response.ok:
                    raise Exception(
                        f"Uploading attachment to SynBioHub failed "
                        f"({response.status_code}): {response.text}"
                    )

                print(
                    f"Uploaded attachment {os.path.basename(sbml_path)} "
                    f"({response.status_code})"
                )
                #print(f"{subCollection_url}/attach")
                #print(response)
                #display_id = make_identifier(sbml_file.filename)
                #model = sbol2.model.Model(uri=display_id)
                #model.source = sbml_file.filename
                #model.language = "http://identifiers.org/edam/format_2585"
                #model.framework = "http://identifiers.org/biomodels.sbo/SBO:0000062"
                #doc.addModel(model)
        subCollection = sbol2.Collection(importType)
        search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,subCollection_url)
        for binding in search_result["results"]["bindings"]:
            uri = binding["s"]["value"]
            subCollection.members = subCollection.members + [ uri ]
        for tl in doc:
            subCollection.members = subCollection.members + [ tl.identity ]
        doc.addCollection(subCollection)
        doc.write(sbol_out_path)
        #print(doc.writeString())

        print("uploading to SBH")
        with open(sbol_out_path, "rb") as f:
            response = requests.post(
                f"{sbh_url}/submit",
                headers={"Accept": "text/plain",
                         "X-authorization": sbh_token,
                },
                files={
                    "files": f,
                },
                data={
                    "rootCollections": sbh_collection_url,
                    "overwrite_merge": sbh_overwrite,
                },
            )
        if not response.ok:
            raise Exception(
                f"SynBioHub submit failed ({response.status_code}): {response.text}"
            )
        return sbh_collection_url
    except AttributeError as e:
        print('Attribute Error: ',str(e))
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": str(e),
            "type": type(e).__name__,
            "repr": repr(e)
        }), 500
    finally:
        if sbol_path is not None and os.path.exists(sbol_path):
            for path in (sbol_path, sbol_out_path):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception as cleanup_error:
                    print(f"Warning: failed to remove temporary file {path}: {cleanup_error}")
        elif sbml_path is not None and os.path.exists(sbml_path):
            for path in (sbml_path, sbol_out_path):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception as cleanup_error:
                    print(f"Warning: failed to remove temporary file {path}: {cleanup_error}")

def _convert_to_sbol(self, sbol_version=2):
    print("converting to SBOL")
    try:
        sbol2.Config.setOption(sbol2.ConfigOptions.SBOL_COMPLIANT_URIS, True)
        sbol2.Config.setOption(sbol2.ConfigOptions.SBOL_TYPED_URIS, False)
        excel2sbol.converter(file_path_in = self.input_excel_path, 
                        file_path_out = self.file_path_out, homespace=self.homespace, sbol_version=sbol_version)
        doc = sbol2.Document()
        doc.read(self.file_path_out)
        print("conversion complete")
        self.sbol_doc = doc
    except Exception as e:
        print("CONVERSION FAILED --- SEE MESSAGE")
        print(f"{type(e).__name__}: {e}")

        # Print full traceback so package-level failures are visible.
        traceback.print_exc()

        # If present, print chained exceptions explicitly for deeper root-cause debugging.
        if e.__cause__ is not None:
            print("\nDirect cause:")
            print(f"{type(e.__cause__).__name__}: {e.__cause__}")
            print("".join(traceback.format_exception(type(e.__cause__), e.__cause__, e.__cause__.__traceback__)))

        if e.__context__ is not None and e.__context__ is not e.__cause__:
            print("\nContext:")
            print(f"{type(e.__context__).__name__}: {e.__context__}")
            print("".join(traceback.format_exception(type(e.__context__), e.__context__, e.__context__.__traceback__)))
        raise

'''
Helper function to upload to SynBioHub and Flapjack using XDC/XDE
'''
def sbh_fj_upload(files):
    if 'Metadata' not in files:
        return 'No file part', 400
    metadata_file = files['Metadata']
    if metadata_file.filename == '':
        return 'No selected file', 400
    root, extension = os.path.splitext(metadata_file.filename)
    if not extension == '.xlsx' and not extension == '.xlsm':
        return 'Invalid Metadata file format', 400

    # Check params from frontend
    if 'Params' not in files:
        return 'No Params file part', 400
    params_file = files['Params']
    if params_file.filename == '':
        return 'No selected Params file', 400
    params_from_request = json.loads(params_file.read())
    sbh_url = params_from_request.get('sbh_url')
    if sbh_url and not (sbh_url.startswith('http://') or sbh_url.startswith('https://')):
        params_from_request['sbh_url'] = 'https://' + sbh_url

    required_params = ['sbh_url', 'sbh_token', 'sbh_user', 'sbh_pass',
                       'collection_url', 'sbh_overwrite']

    for param in required_params:
        if param not in params_from_request:
            return 'Parameter ' + param + ' not found in request', 400
    if (params_from_request['sbh_token'] is None and 
        params_from_request['sbh_user'] is None and
        params_from_request['sbh_pass'] is None):
        return 'No SBH credentials provided', 400
    
    fj_url = params_from_request.get('fj_url')
    fj_token = params_from_request.get('fj_token')
    fj_user = params_from_request.get('fj_user')
    fj_pass = params_from_request.get('fj_pass')
    fj_overwrite = params_from_request.get('fj_overwrite', 1)

    if not fj_url:
        fj_url = None
        fj_token = None
        fj_user = None
        fj_pass = None
    elif not fj_token and not (fj_user and fj_pass):
        return jsonify({
            "error": "Flapjack URL was provided, but no Flapjack credentials were provided"
        }), 400

    # Attachment files to upload to SBH
    if 'Attachments' in files and 'attachments' in params_from_request:
        attachment_files = files.getlist("Attachments")
        attachments = {}
        for file in attachment_files:
            if file.filename not in params_from_request['attachments']:
                return (
                    f"Attachment metadata for file '{file.filename}' not found in request",
                    400,
                )
            attachments[params_from_request['attachments'][file.filename]] = file
    else:
        attachments = None

    upload_dir = os.path.join(os.getcwd(), "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    safe_metadata_filename = secure_filename(metadata_file.filename)

    if safe_metadata_filename == '':
        return 'Invalid Metadata file name', 400

    metadata_path = os.path.join(
        upload_dir,
        f"{uuid4()}_{safe_metadata_filename}"
    )

    metadata_file.save(metadata_path)

    # Plate reader data to upload to FJ
    if 'Plate_Reader_Output' in request.files and 'sheet_name' in params_from_request:
        filenames = [metadata_path]
        for file in files.getlist("Plate_Reader_Output"):
            # TODO - adapt XDE to work with the file object to avoid unneccesary writes
            safe_data_filename = secure_filename(file.filename)
            if safe_data_filename == '':
                return 'Invalid Plate Reader Output file name', 400
            data_path = os.path.join(upload_dir, safe_data_filename)
            file.save(data_path)
            filenames.append(data_path)
        xde = tricahue.XDE()
        xde.run(filenames, params_from_request['sheet_name'], data_cols_offset=2)
        print(filenames)
        for data_filename in filenames[1:]:
            os.remove(data_filename)

    # instantiate the XDC class using the params_from_request dictionary
    try:
        xdc = tricahue.XDC(input_excel_path = metadata_path, attachments=attachments)
        # print(params_from_request['sbh_url'], params_from_request['collection_url'], params_from_request['sbh_overwrite'], params_from_request['sbh_user'],params_from_request['sbh_pass'], params_from_request['sbh_pass'],params_from_request['fj_url'], params_from_request['fj_overwrite'], params_from_request['fj_user'], params_from_request['fj_pass'],params_from_request['fj_token'])
        sbh_url, fj_url = xdc.upload_to_existing_collection(sbh_url = params_from_request['sbh_url'],
                                      collection_url = params_from_request['collection_url'], 
                                      sbh_overwrite = params_from_request['sbh_overwrite'], 
                                      sbh_user = params_from_request['sbh_user'],
                                      sbh_pass = params_from_request['sbh_pass'], 
                                      sbh_token = params_from_request['sbh_token'],
                                      fj_url = fj_url,
                                      fj_overwrite = fj_overwrite, 
                                      fj_user = fj_user, 
                                      fj_pass = fj_pass,
                                      fj_token = fj_token,
                                      importType = params_from_request['importType'])
    except AttributeError as e:
        os.remove(metadata_path)
        print('Attribute Error: ',str(e))
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback

        traceback.print_exc()

        return jsonify({
            "error": str(e),
            "type": type(e).__name__,
            "repr": repr(e)
        }), 500
    
    sbs_upload_response_dict ={
        "sbh_url": sbh_url,
        "fj_url": fj_url,
        "status": "success"
    }
    os.remove(metadata_path)
    return jsonify(sbs_upload_response_dict)


@app.route('/sbol_2_build_golden_gate', methods=['POST'])
def sbol_2_build_golden_gate():
    # Error checking in the request
    print("request", request.files)

    if 'plasmid_backbone' not in request.files:
        return jsonify({"error": "Missing plasmid backbone"}), 400
    if 'insert_parts' not in request.files:
        return jsonify({"error": "Missing insert parts"}), 400
    if 'wizard_selections' not in request.form:
        return jsonify({"error": "Missing wizard selections"}), 400

    wizard_selection = request.form.get('wizard_selections')
    plasmid_backbone = request.files.get('plasmid_backbone')
    insert_parts = request.files.getlist('insert_parts')

    # Parse the json

    wizard_selection_json = json.loads(wizard_selection)
    assembly_method = wizard_selection_json.get('formValues').get('assemblyMethod')

    # Check if the assembly method is valid
    if assembly_method != 'MoClo':
        return jsonify({"error": "Invalid assembly method"}), 400
    
    # Get the restriction item
    restriction_enzyme = wizard_selection_json.get('formValues').get('restrictionEnzyme')

    # code for sbol2build
    part_docs = []
    for item in insert_parts:
        doc = sb2.Document()
        doc.read(item)
        part_docs.append(doc)
    
    bb_doc = sb2.Document()
    bb_doc.read(plasmid_backbone)

    assembly_doc = sb2.Document()
    assembly_obj = sbol2build.golden_gate_assembly_plan('testassem', part_docs, bb_doc, restriction_enzyme, assembly_doc)

    try:
        # Run abstract translator to get plasmids
        plasmid_documents, vector_doc, design_id = abstract_design_2_plasmids(abstract_design_uri, plasmid_collection_uri, plasmid_vector_uri, sbh)
        
        # Run plasmids through sbol2build to generate assembly plan
        assembly_plan_doc = sbol2build_moclo(plasmid_documents, vector_doc, design_id)
        assembly_plan_doc.displayId = f"{design_id}_assembly"

    
        sbh_response = sbh.submit(
            doc=assembly_plan_doc,
            collection=recipient_collection_uri,
            overwrite=2
        )
        return sbh_response.text, sbh_response.status_code

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        return jsonify({"error": f"Unexpected server error: {str(e)}"}), 500
    
"""@app.route('/upload_transformation', methods=['POST'])    
def upload_transformation():
    if 'auth_token' not in request.form:
        return jsonify({"error": "Missing SynBioHub Authentication Token"}), 400
    if 'registry_url' not in request.form:
        return jsonify({"error": "Missing SynBioHub Registry URL"}), 400
    if 'collection_uri' not in request.form:
        return jsonify({"error": "Missing recipient SynBioHub collection URI"}), 400

    if 'plasmid_uris' not in request.form:
        return jsonify({"error": "Missing plasmid URIs"}), 400
    if 'chassis_uri' not in request.form:
        return jsonify({"error": "Missing chassis URI"}), 400
    if 'machine' not in request.form:
        return jsonify({"error": "Missing machine"}), 400
    if 'protocol' not in request.form:
        return jsonify({"error": "Missing protocol"}), 400
    if 'params' not in request.form:
        return jsonify({"error": "Missing trasnformation parameters"}), 400

    auth_token = request.form.get("auth_token")
    sbh_registry = request.form.get("registry_url")
    recipient_collection_uri = request.form.get("collection_uri")

    plasmid_uris = request.form.get("plasmid_uris")
    chassis_uri = request.form.get("chassis_uri")
    machine_name = request.form.get("machine")
    protocol = request.form.get("protocol")
    parameters = request.form.get("params")

    sbh = sbol2.PartShop(sbh_registry)
    sbh.key = auth_token
    
    try:
        transformation_doc = generate_transformation_metadata(plasmid_uris, chassis_uri, machine_name, protocol, parameters, sbh)

        sbh_response = sbh.submit(
            doc=transformation_doc,
            collection=recipient_collection_uri,
            overwrite=2
        )
        return sbh_response.text, sbh_response.status_code

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        return jsonify({"error": f"Unexpected server error: {str(e)}"}), 500"""

@app.route('/api/build_pudu', methods=['POST'])
def build_pudu():
    # Error checking in the request
    print("request", request.files)

    if 'assembly_plan' not in request.files:
        return jsonify({"error": "Missing assembly plan"}), 400
    if 'wizard_selections' not in request.form:
        return jsonify({"error": "Missing wizard selections"}), 400

    wizard_selection = request.form.get('wizard_selections')
    assembly_plan_file = request.files.get('assembly_plan')

    # # Parse the json
    wizard_selection_json = json.loads(wizard_selection)
    build_method = wizard_selection_json.get('formValues').get('buildMethod')

    # Check if the assembly method is valid
    if build_method != 'PUDU':
        return jsonify({"error": "Invalid build method"}), 400

    # TODO: save xml to a file ('assembly_plan.xml')

    try:
        # Run script (which has opentrons script hardcoded) using JSON file
        log = subprocess.run(["python", "run_sbol2assembly.py"], capture_output=True).stdout
        curpath = os.path.abspath(os.curdir)
        print(curpath)
        # write captured output to a text file
        # w = write mode, create file if doesn't exist; b = binary file
        with open("files/build_log.txt", "wb") as log_file:
            log_file.write(log)
        # read excel file "sbol2_assembly_output.xlsx"

        # returns: build_log.txt, excel file, py protocol, build plan
        return jsonify({"message": "PUDU build not implemented yet"}), 501
    
    except ValueError as e:
        # catch errors and return to frontend
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Unexpected server error: {str(e)}"}), 500


@app.route('/api/inspect_request', methods=['POST'])
def inspect_request():
    files = {}
    for name in request.files:
        file = request.files[name]
        try:
            files[name] = json.loads(file.read())
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    return jsonify({
        "message": "Request received successfully", 
        "files": files}), 200

