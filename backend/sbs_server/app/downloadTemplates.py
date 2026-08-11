from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from flask import jsonify, send_file
import requests
from .utils import sbh_get_subCollection_uris
import json
from io import BytesIO

'''
Helper function to download Excel template
'''
def fetch_template_bytes(object_type_id):
    templates = {
        "synbio.object-type.sample-designs": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/SampleDesign.xlsm",
            "SampleDesign.xlsm",
        ),
        "synbio.object-type.strains": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Strains.xlsm",
            "Strains.xlsm",
        ),
        "synbio.object-type.resources": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Resources.xlsm",
            "Resources.xlsm",
        ),
        "synbio.object-type.study-data": (
            "https://raw.githubusercontent.com/SynBioDex/Excel-to-SBOL/master/resources/templates/Assay.xlsm",
            "Assay.xlsm",
        ),
    }

    if object_type_id not in templates:
        raise ValueError(f"Unknown object type: {object_type_id}")

    url, filename = templates[object_type_id]
    response = requests.get(url)
    response.raise_for_status()
    return response.content, filename

def expand_table_in_xlsm(wb, sheet_name, table_name, new_rows):
    ws = wb[sheet_name]
    table = ws.tables[table_name]

    start_cell, end_cell = table.ref.split(":")
    start_col_letters = "".join(ch for ch in start_cell if ch.isalpha())
    start_row = int("".join(ch for ch in start_cell if ch.isdigit()))
    end_col_letters = "".join(ch for ch in end_cell if ch.isalpha())
    end_row = int("".join(ch for ch in end_cell if ch.isdigit()))

    start_col = column_index_from_string(start_col_letters)
    end_col = column_index_from_string(end_col_letters)

    # Find the first blank row inside the existing table body
    first_blank_row = None
    for row_idx in range(start_row + 1, end_row + 1):
        if all(ws.cell(row=row_idx, column=col_idx).value is None for col_idx in range(start_col, end_col + 1)):
            first_blank_row = row_idx
            break

    # Write into the blank row first if one exists
    write_row = first_blank_row if first_blank_row is not None else end_row + 1

    for values in new_rows:
        # If we run past the current table range, extend it
        if write_row > end_row:
            end_row = write_row
            table.ref = f"{start_cell}:{end_col_letters}{end_row}"

        for col_idx, value in enumerate(values, start=start_col):
            ws.cell(row=write_row, column=col_idx, value=value)

        write_row += 1

'''
Helper function to download Excel template and populate tables from SynBioHub
'''
def sbh_download_template(files):
    if 'Params' not in files:
        return jsonify({"error": "No Params file part"}), 400

    params_file = files['Params']
    if params_file.filename == '':
        return jsonify({"error": "No selected Params file"}), 400

    params_from_request = json.loads(params_file.read())

    required_params = ['template_type', 'sbh_url', 'sbh_token', 'collection_url']
    for param in required_params:
        if param not in params_from_request:
            return jsonify({"error": f"Parameter {param} not found in request"}), 400

    if params_from_request['sbh_token'] is None:
        return jsonify({"error": "No SBH credentials provided"}), 400

    sbh_url = params_from_request['sbh_url']
    sbh_collection_url = params_from_request['collection_url'] 
    sbh_token = params_from_request['sbh_token']
    parts = sbh_collection_url.split("/")
    usergraph = "/".join(parts[:5])

    template_type = params_from_request['template_type']

    try:
        template_bytes, filename = fetch_template_bytes(template_type)
        
        if template_type == "synbio.object-type.resources":
            return send_file(
                BytesIO(template_bytes),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
            )

        wb = load_workbook(BytesIO(template_bytes), keep_vba=True)

        if template_type == "synbio.object-type.strains":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/ncit/NCIT:C14419")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_chassis_collections",table_name="SBH_chassis_collections",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/so/SO:0000637")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_plasmids_collections",table_name="SBH_plasmids_collections",new_rows=new_rows)
        elif template_type == "synbio.object-type.sample-designs":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://identifiers.org/ncit/NCIT:C48164")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_media_collection",table_name="SBH_media_collection",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,type="http://www.biopax.org/release/biopax-level3.owl#SmallMolecule")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_chemicals_collection",table_name="SBH_chemicals_collection",new_rows=new_rows)
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"http://purl.obolibrary.org/obo/NCIT_C97158")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_strains_collection",table_name="SBH_strains_collection",new_rows=new_rows)
        elif template_type == "synbio.object-type.study-data":
            search_result = sbh_get_subCollection_uris(sbh_url,sbh_token,usergraph,sbh_collection_url,"https://wiki.synbiohub.org/wiki/Terms/SynBioSuite#SampleDesign")
            new_rows = []
            for binding in search_result["results"]["bindings"]:
                uri = binding["s"]["value"]
                id = binding["id"]["value"]
                new_rows.append([id, uri])
            expand_table_in_xlsm(wb,sheet_name="SBH_sampledesigns_collection",table_name="SBH_sampledesigns_collection",new_rows=new_rows)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500
