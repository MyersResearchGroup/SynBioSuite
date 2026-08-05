from __future__ import annotations

import asyncio
import csv
import datetime
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import sbol2
import tricahue
import websockets
from flapjack import Flapjack
from openpyxl import Workbook, load_workbook

# NCIT role codes that classify a sample-design module. The role URIs aren't
# uniform (medium uses ``NCIT:C48164``, strain uses ``NCIT_C97158``), so callers
# match with ``endswith`` rather than splitting on a separator.
MEDIUM_ROLE = "C48164"
STRAIN_ROLES = "C14419"  #or bare chassis (C14419)

# 96-well plate layout: rows A-H, cols 1-12, wells in A1..A12,B1..H12 order.
PLATE_ROWS = tuple("ABCDEFGH")
PLATE_COLS = tuple(range(1, 13))
WELLS = ["{0}{1}".format(row, col) for row in PLATE_ROWS for col in PLATE_COLS]


# --------------------------------------------------------------------------- #
# SynBioHub client (PartShop)
# --------------------------------------------------------------------------- #
def get_sbh_client(
    email: Optional[str] = None,
    password: Optional[str] = None,
    *,
    token: Optional[str] = None,
    url: str = "https://synbiohub.org",
) -> sbol2.PartShop:
    """Return an authenticated PartShop -- the SBH client.

    Pass a SynBioHub ``token`` (the intended workflow -- pySBOL2 sends it as the
    ``X-authorization`` header), or fall back to ``email``/``password``. The only
    function that takes SynBioHub credentials; :func:`resolve` reuses the returned
    PartShop and never needs auth again.
    """
    shop = sbol2.PartShop(url)
    if token:
        shop.key = token
    elif email and password:
        shop.login(email, password)
    else:
        raise ValueError("get_sbh_client needs a token, or an email + password")
    return shop


# --------------------------------------------------------------------------- #
# The one design-domain object type + resolving it from SBOL
# --------------------------------------------------------------------------- #
@dataclass
class Supplement:
    """A chemical added to a sample design, with its concentration."""

    chemical: str  # name
    uri: str
    concentration: float


@dataclass
class SampleDesign:
    """A sample design resolved from SBOL: media, strain, plasmids, supplements.

    ``media``/``strain`` are ``(name, uri)``; ``strain`` is the *engineered*
    strain (role C97158) whose plasmids become Flapjack DNA. Each field is enough
    to get-or-create the matching Flapjack object.
    """

    id: str  # the design name
    uri: str
    media: Optional[Tuple[str, str]] = None
    strain: Optional[Tuple[str, str]] = None
    plasmids: List[Tuple[str, str]] = field(default_factory=list)
    supplements: List[Supplement] = field(default_factory=list)


def _name_from_uri(uri: str) -> str:
    """displayId at the end of a SynBioHub URI (handles a trailing ``/1`` version)."""
    parts = uri.rstrip("/").split("/")
    return parts[-2] if parts[-1].isdigit() else parts[-1]


def _has_role(module_definition, codes) -> bool:
    """True if any of the module's roles ends with one of ``codes``."""
    codes = (codes,) if isinstance(codes, str) else codes
    return any(
        str(role).rstrip("/").endswith(code)
        for role in module_definition.roles
        for code in codes
    )


def _concentration(module_definition) -> Optional[float]:
    """The Flapjack#concentration value if the module carries one, else None."""
    for key, values in module_definition.properties.items():
        if str(key).endswith("concentration") and values:
            return float(values[0])
    return None


def resolve(shop: sbol2.PartShop, name: str, uri: str) -> SampleDesign:
    """Get a sample design from SBH and resolve it into a :class:`SampleDesign`.

    One recursive pull, then a purely local walk: classify each module of the
    design as medium (role C48164), strain (C97158/C14419, whose functional
    components are the plasmids), or supplement (no role but a concentration,
    whose functional component is the chemical).
    """
    # one recursive pull brings the whole design tree (design + medium + strain +
    # chassis) in a single round-trip; plasmids/chemicals ride as URI references
    document = sbol2.Document()
    shop.pull(uri, document, recursive=True)
    by_uri = {md.identity: md for md in document.moduleDefinitions}
    design = SampleDesign(id=name, uri=uri)
    for module in by_uri[uri].modules:
        sub = by_uri[str(module.definition)]
        if _has_role(sub, MEDIUM_ROLE):
            design.media = (sub.displayId, sub.identity)
        elif _has_role(sub, STRAIN_ROLES):
            design.strain = (sub.displayId, sub.identity)
            design.plasmids = [
                (_name_from_uri(str(fc.definition)), str(fc.definition))
                for fc in sub.functionalComponents
            ]
        elif (concentration := _concentration(sub)) is not None:
            for fc in sub.functionalComponents:
                uri_ = str(fc.definition)
                design.supplements.append(Supplement(_name_from_uri(uri_), uri_, concentration))
        else:
            raise ValueError(
                "Cannot classify module {0!r} of design {1}".format(sub.displayId, name)
            )
    return design


# --------------------------------------------------------------------------- #
# Flapjack client + get-or-create
# --------------------------------------------------------------------------- #
def get_flapjack_client(
    username: str,
    password: Optional[str] = None,
    *,
    access_token: Optional[str] = None,
    refresh_token: Optional[str] = None,
    url: str = "charmmefj-api.synbiohub.org",
) -> Flapjack:
    """Return an authenticated Flapjack client (the API host, not the web host).

    Pass an ``access_token`` (+ ``refresh_token``) -- the intended workflow, via
    Flapjack's ``log_in_token`` -- or fall back to ``password``. Flapjack tokens
    are JWTs; the only way to mint one is Flapjack's own ``/api/auth/log_in/``.
    """
    flapjack = Flapjack(url_base=url)
    if access_token:
        flapjack.log_in_token(username, access_token, refresh_token)
    elif password:
        flapjack.log_in(username=username, password=password)
    else:
        raise ValueError("get_flapjack_client needs an access_token, or a password")
    return flapjack


def get_or_create(flapjack: Flapjack, model: str, match: dict, fields: dict) -> int:
    """Return a Flapjack object's id, creating it only if absent.

    ``match`` is the identity filter for the existence check -- the identity key
    *only* (``name`` for media/strain/study, ``sboluri`` for dna/chemical), never
    the full body, or an object with a differing description would be missed and
    duplicated. ``fields`` is the body POSTed when the object doesn't exist.
    """
    existing = flapjack.get(model, **match)
    if len(existing):
        return int(existing.id.values[0])
    created = flapjack.create(model, confirm=False, overwrite=False, **fields)
    return int(created.id.values[0])


def provision(flapjack: Flapjack, design: SampleDesign) -> Dict[str, object]:
    """Get-or-create every Flapjack object a :class:`SampleDesign` references.

    Returns ``{'media': id|None, 'strain': id|None, 'dna': [ids], 'chemical': [ids]}``.
    Media/strain match on name; DNA/chemical match on sboluri.
    """
    ids: Dict[str, object] = {"media": None, "strain": None, "dna": [], "chemical": []}
    if design.media:
        name, uri = design.media
        ids["media"] = get_or_create(flapjack, "media", {"name": name},
                                     {"name": name, "description": name, "sboluri": uri})
    if design.strain:
        name, uri = design.strain
        ids["strain"] = get_or_create(flapjack, "strain", {"name": name},
                                      {"name": name, "description": name, "sboluri": uri})
    for name, uri in design.plasmids:
        ids["dna"].append(get_or_create(flapjack, "dna", {"sboluri": uri}, {"name": name, "sboluri": uri}))
    for supplement in design.supplements:
        ids["chemical"].append(
            get_or_create(
                flapjack, "chemical", {"sboluri": supplement.uri},
                {"name": supplement.chemical, "description": supplement.chemical, "sboluri": supplement.uri},
            )
        )
    return ids


# --------------------------------------------------------------------------- #
# Read the study workbook's metadata
# --------------------------------------------------------------------------- #
def _find_sheet(workbook, name: str):
    """Case-insensitive worksheet lookup."""
    try:
        return next(w for w in (workbook[n] for n in workbook.sheetnames) if w.title.lower() == name.lower())
    except StopIteration:
        raise ValueError("workbook has no {0!r} sheet".format(name))


def _header_index(sheet) -> Dict[str, int]:
    """``{normalized column name -> column index}`` from the sheet's first row."""
    return {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(next(sheet.iter_rows()))
        if cell.value is not None
    }


def _read_well_designs(
    workbook_path: Union[str, Path],
    assay_id: Optional[str] = None,
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """``({well -> design name}, {design name -> URI})`` for an assay's wells.

    The single reader for the ``sample`` + ``SBH_sampledesigns_collection`` sheets.
    Skips blank rows and validates each well is inside the 96-well plate.
    """
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        collection = _find_sheet(workbook, "SBH_sampledesigns_collection")
        cols = _header_index(collection)
        uris = {
            str(r[cols["name"]]): str(r[cols["uri"]])
            for r in collection.iter_rows(min_row=2, values_only=True)
            if r[cols["name"]] and r[cols["uri"]]
        }
        samples = _find_sheet(workbook, "sample")
        scols = _header_index(samples)
        well_design: Dict[str, str] = {}
        for r in samples.iter_rows(min_row=2, values_only=True):
            if assay_id is not None and r[scols["assay id"]] != assay_id:
                continue
            if not (r[scols["row"]] and r[scols["column"]] and r[scols["sample design"]]):
                continue
            row_name, column = str(r[scols["row"]]).upper(), int(r[scols["column"]])
            if row_name not in PLATE_ROWS or column not in PLATE_COLS:
                raise ValueError(
                    "sample well {0}{1} is outside the 96-well plate (rows A-H, cols 1-12)"
                    .format(row_name, column)
                )
            well_design["{0}{1}".format(row_name, column)] = str(r[scols["sample design"]])
    finally:
        workbook.close()
    return well_design, uris


def read_unique_designs(
    workbook_path: Union[str, Path],
    assay_id: Optional[str] = None,
) -> List[Tuple[str, str]]:
    """The distinct ``(design name, SynBioHub URI)`` used by an assay, first-seen order.

    With ``assay_id`` given, only that assay's wells count; otherwise all wells.
    Raises if a referenced design is missing from ``SBH_sampledesigns_collection``.
    """
    well_design, uris = _read_well_designs(workbook_path, assay_id)
    names = list(dict.fromkeys(well_design.values()))
    missing = [n for n in names if n not in uris]
    if missing:
        raise ValueError("designs missing from SBH_sampledesigns_collection: {0}".format(missing))
    return [(name, uris[name]) for name in names]


def resolve_designs(
    shop: sbol2.PartShop,
    workbook_path: Union[str, Path],
    assay_id: Optional[str] = None,
) -> Dict[str, SampleDesign]:
    """``{design name -> SampleDesign}`` for every unique design used by an assay."""
    return {name: resolve(shop, name, uri)
            for name, uri in read_unique_designs(workbook_path, assay_id)}


def read_signals(workbook_path: Union[str, Path]) -> List[Tuple[str, str, str]]:
    """Ordered ``[(name, description, color)]`` from the workbook's ``signal`` sheet.

    Row order matters: it is matched to the reader's read order to build the
    reader-channel -> signal-name mapping (see :func:`import_study`). Description
    falls back to the name, color to a neutral gray.
    """
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = _find_sheet(workbook, "signal")
        cols = _header_index(sheet)
        name_i, desc_i, color_i = (cols.get("signal name"), cols.get("signal description"),
                                   cols.get("signal color"))
        signals: List[Tuple[str, str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[name_i] if name_i is not None else None
            if not name:
                continue
            description = str(row[desc_i]) if desc_i is not None and row[desc_i] else str(name)
            color = str(row[color_i]) if color_i is not None and row[color_i] else "#888888"
            signals.append((str(name), description, color))
    finally:
        workbook.close()
    return signals


# --------------------------------------------------------------------------- #
# Construct WB -- build the Flapjack input template (no Flapjack writes)
# --------------------------------------------------------------------------- #
_INT_RE = re.compile(r"[+-]?\d+")
_NUM_RE = re.compile(r"[+-]?(?:\d+\.\d*|\d*\.\d+|\d+)(?:[eE][+-]?\d+)?")


def _excel_value(text):
    """A raw plate-file cell string -> int / float / str / None."""
    text = (text or "").strip()
    if not text:
        return None
    if _INT_RE.fullmatch(text):
        return int(text)
    if _NUM_RE.fullmatch(text):
        return float(text)
    return text


def _hours_to_time(hours) -> datetime.time:
    """A float number of hours -> datetime.time (Flapjack's parser reads .hour/.minute)."""
    total = round(float(hours) * 3600)
    return datetime.time(total // 3600, (total % 3600) // 60, total % 60)


def _excel_number(value):
    try:
        number = float(value)
        return int(number) if number == int(number) else number
    except (TypeError, ValueError):
        return value


def _write_grid(sheet, title, well_values, start_row: int = 1, blank_missing: bool = False):
    """Write one labelled 96-well grid.

    A missing well becomes the string ``'None'`` (what the reference files use for
    Media/Strains/DNA), or a blank cell when ``blank_missing`` -- for Chemicals,
    since the server reads a blank chemical cell as 0 but crashes on ``float('None')``.
    """
    sheet.cell(start_row, 1, title)
    for col in PLATE_COLS:
        sheet.cell(start_row, col + 1, col)
    for i, row_name in enumerate(PLATE_ROWS, start=start_row + 1):
        sheet.cell(i, 1, row_name)
    for well, value in well_values.items():
        row = start_row + PLATE_ROWS.index(well[0]) + 1
        col = int(well[1:]) + 1
        if value is None:
            if blank_missing:
                continue
            value = "None"
        sheet.cell(row, col, value)


def construct_wb(
    workbook_path: Union[str, Path],
    plate_file: Union[str, Path],
    designs: Dict[str, SampleDesign],
    output: Union[str, Path],
    *,
    block_names: Dict[str, str],
    assay_id: Optional[str] = None,
    temperature: int = 37,
) -> Path:
    """Build the Flapjack input template and return its path -- writes nothing to Flapjack.

    ``designs`` is ``{design name -> SampleDesign}`` (see :func:`resolve_designs`).
    ``block_names`` maps each reader Signal ID (e.g. ``"Read 1:569,593"``) to a
    clean, comma-free Flapjack signal name, in signal order. The Data sheet is the
    plate file's preamble through ``End Kinetic``, then a Time/Temp/well block per
    signal (from ``read_neo`` on the same file), then a ``Results`` sentinel; the
    Media/Strains/DNA/Chemicals sheets are 96-well grids from ``designs``.
    """
    well_design, _ = _read_well_designs(workbook_path, assay_id)

    def per_well(pick):
        return {well: pick(designs[name]) for well, name in well_design.items()}

    workbook = Workbook()

    # Data sheet: preamble from the plate file, through 'End Kinetic'
    data = workbook.active
    data.title = "Data"
    with open(plate_file, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = [[None]] + [[_excel_value(v) for v in row] for row in csv.reader(handle)]
    end = next((i for i, r in enumerate(rows) if r and r[0] == "End Kinetic"), None)
    if end is None:
        raise ValueError("'End Kinetic' marker not found in plate file: {0}".format(plate_file))
    for row in rows[:end + 1]:
        data.append(row)
    data.append([])

    # Data sheet: one block per signal, from read_neo on the same file
    frame = tricahue.XDE().read_neo(str(plate_file))
    for reader_signal, clean_name in block_names.items():
        wide = (
            frame[frame["Signal ID"] == reader_signal]
            .pivot_table(index="Time", columns="Sample ID", values="Value", aggfunc="first")
            .reindex(columns=WELLS)
        )
        data.append([clean_name])
        data.append([])
        data.append([None, "Time", "T° {0}".format(clean_name)] + list(WELLS))
        for time in wide.index:
            data.append([None, _hours_to_time(time), temperature]
                        + [_excel_number(wide.loc[time, well]) for well in WELLS])
        data.append([])
    data.append(["Results"])
    data.append([])

    # metadata grids
    _write_grid(workbook.create_sheet("Media"), "Media",
                per_well(lambda d: d.media[0] if d.media else None))
    _write_grid(workbook.create_sheet("Strains"), "Strains",
                per_well(lambda d: d.strain[0] if d.strain else None))

    dna = workbook.create_sheet("DNA")
    n_plasmids = max((len(d.plasmids) for d in designs.values()), default=0)
    for i in range(n_plasmids):
        _write_grid(dna, "DNA {0}".format(i + 1),
                    per_well(lambda d, i=i: d.plasmids[i][0] if i < len(d.plasmids) else None),
                    start_row=i * 10 + 1)

    chemical_names = sorted({s.chemical for d in designs.values() for s in d.supplements})
    if chemical_names:                       # only add a Chemicals sheet if a design has one
        chem = workbook.create_sheet("Chemicals")
        for i, name in enumerate(chemical_names):
            _write_grid(
                chem, name,
                per_well(lambda d, c=name: next((s.concentration for s in d.supplements if s.chemical == c), None)),
                start_row=i * 10 + 1,
                blank_missing=True,          # empty chemical wells -> blank (server reads 0)
            )

    output = Path(output)
    workbook.save(output)
    return output


# --------------------------------------------------------------------------- #
# Upload WB -- live-writing step (websocket file upload)
# --------------------------------------------------------------------------- #
async def _upload_wb(flapjack, wb_file, study_id, assay_name, machine, temperature,
                     description, signal_ids, dna_ids, chemical_ids):
    flapjack.refresh()                                   # a fresh access token for the ws url
    file_bytes = Path(wb_file).read_bytes()
    form = {
        "study": study_id, "name": assay_name, "machine": machine,
        "description": description or assay_name, "temperature": temperature,
    }
    uri = (flapjack.ws_url_base.replace("ws://", "wss://")
           + "/ws/registry/upload?token=" + flapjack.access_token)
    assay_id = None
    async with websockets.connect(uri, max_size=int(1e10)) as socket:
        await socket.send(json.dumps({"type": "init_upload", "data": form}))
        try:
            while True:
                message = json.loads(await asyncio.wait_for(socket.recv(), timeout=600))
                kind = message.get("type")
                if kind == "ready_for_file":
                    assay_id = message["data"]["assay_id"]       # init_upload created the Assay
                    await socket.send(file_bytes)
                elif kind == "input_requests":
                    requested = message["data"]                  # names the server parsed from the grids
                    # every requested name must have a provisioned id, else the server
                    # would 500 mid-creation and orphan the assay -- fail fast instead.
                    missing = ([n for n in requested.get("signal", []) if n not in signal_ids]
                               + [n for n in requested.get("dna", []) if n not in dna_ids]
                               + [n for n in requested.get("chemical", []) if n not in chemical_ids])
                    if missing:
                        raise RuntimeError("no provisioned Flapjack id for: {0}".format(missing))
                    payload = {
                        "signal": [signal_ids[n] for n in requested.get("signal", [])],
                        "dna": [dna_ids[n] for n in requested.get("dna", [])],
                        "chemical": [chemical_ids[n] for n in requested.get("chemical", [])],
                    }
                    await socket.send(json.dumps({"type": "metadata", "data": payload}))
                elif kind == "creation_done":
                    return assay_id
                elif kind == "error" or (kind and "error" in str(kind)):
                    raise RuntimeError("Flapjack upload error: " + json.dumps(message)[:300])
        except Exception:
            if assay_id is not None:                     # don't leave a half-created assay behind
                try:
                    flapjack.delete("assay", assay_id, confirm=False)
                except Exception:
                    pass
            raise


def upload_wb(
    flapjack: Flapjack,
    wb_file: Union[str, Path],
    study_id: int,
    assay_name: str,
    *,
    signal_ids: Dict[str, int],
    dna_ids: Dict[str, int],
    chemical_ids: Optional[Dict[str, int]] = None,
    machine: str = "HTX Synergy",
    temperature: int = 37,
    description: Optional[str] = None,
) -> int:
    """Upload a constructed workbook and return the new assay id -- the live-writing step.

    ``init_upload`` creates the Assay under ``study_id``; the file streams; when the
    server sends ``input_requests`` (the DNA/Chemical/Signal names it parsed from the
    grids) we answer with our provisioned ids, and it builds Samples + Measurements.
    ``signal_ids``/``dna_ids``/``chemical_ids`` map a name -> Flapjack id. On any
    failure after the assay is created, that assay is deleted so nothing is orphaned.
    (Flapjack applies nest_asyncio, so this runs from both scripts and notebooks.)
    """
    return asyncio.get_event_loop().run_until_complete(
        _upload_wb(flapjack, wb_file, study_id, assay_name, machine, temperature,
                   description, signal_ids, dna_ids, chemical_ids or {})
    )


# --------------------------------------------------------------------------- #
# End-to-end
# --------------------------------------------------------------------------- #
def import_study(
    shop: sbol2.PartShop,
    flapjack: Flapjack,
    workbook_path: Union[str, Path],
    plate_file: Union[str, Path],
    study_id: int,
    assay_name: str,
    output: Union[str, Path],
    *,
    block_names: Optional[Dict[str, str]] = None,
    assay_id: Optional[str] = None,
    machine: str = "HTX Synergy",
    temperature: int = 37,
) -> Dict[str, object]:
    """Run the whole workflow: resolve designs -> provision objects -> construct the
    template -> upload it into ``study_id``.

    Signals are read from the workbook's ``signal`` sheet (name/description/color).
    The reader-channel -> signal-name mapping is matched **by order** (the reader's
    read order against the sheet's row order); pass ``block_names`` to override.

    Returns ``{study_id, assay_id, workbook, dna_ids, chemical_ids, signal_ids}``.
    ``study_id`` must already exist (create it with ``get_or_create(flapjack, "study", ...)``).
    """
    designs = resolve_designs(shop, workbook_path, assay_id)

    # Create SD Objects: provision each design, keeping name -> id maps for the upload
    dna_ids: Dict[str, int] = {}
    chemical_ids: Dict[str, int] = {}
    for design in designs.values():
        provisioned = provision(flapjack, design)
        for (plasmid_name, _), dna_id in zip(design.plasmids, provisioned["dna"]):
            dna_ids[plasmid_name] = dna_id
        for supplement, chemical_id in zip(design.supplements, provisioned["chemical"]):
            chemical_ids[supplement.chemical] = chemical_id

    # Signals come from the workbook's 'signal' sheet; the reader-channel -> name
    # mapping (block_names) is matched by ORDER (reader read order <-> sheet row
    # order) unless block_names was passed explicitly.
    signals = read_signals(workbook_path)
    if block_names is None:
        # reader channels in read order (first appearance in the plate file)
        frame = tricahue.XDE().read_neo(str(plate_file))
        channels = list(dict.fromkeys(frame["Signal ID"]))
        if not signals:
            raise ValueError("No signals: fill the workbook 'signal' sheet, or pass block_names.")
        if len(signals) != len(channels):
            raise ValueError(
                "signal sheet has {0} signal(s) but the plate file has {1} reader channel(s) {2}; "
                "fix the sheet or pass block_names explicitly."
                .format(len(signals), len(channels), channels)
            )
        block_names = {channel: name for channel, (name, _, _) in zip(channels, signals)}
    elif not signals:
        signals = [(name, name, "#888888") for name in dict.fromkeys(block_names.values())]

    signal_ids = {
        name: get_or_create(flapjack, "signal", {"name": name},
                            {"name": name, "description": description, "color": color})
        for name, description, color in signals
    }

    workbook = construct_wb(workbook_path, plate_file, designs, output,
                            block_names=block_names, assay_id=assay_id, temperature=temperature)
    new_assay_id = upload_wb(flapjack, workbook, study_id, assay_name,
                             signal_ids=signal_ids, dna_ids=dna_ids, chemical_ids=chemical_ids,
                             machine=machine, temperature=temperature)
    return {
        "study_id": study_id, "assay_id": new_assay_id, "workbook": workbook,
        "dna_ids": dna_ids, "chemical_ids": chemical_ids, "signal_ids": signal_ids,
    }
